from __future__ import annotations

from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from analysis_common import (
    DATA_CLEAN_RAW,
    DATE_COL,
    EXPECTED_COLUMNS,
    INPUT_FILE,
    LOGS_DIR,
    NUMERIC_COLUMNS,
    SHEET_NAME,
    TABLES_DIR,
    ensure_directories,
    finite_numeric_frame,
    format_date,
    missing_months,
    relative_path,
    save_csv,
    to_month_end,
)


AUDIT_TABLE = TABLES_DIR / "data_audit.csv"
MISSING_TABLE = TABLES_DIR / "missing_values.csv"
AUDIT_LOG = LOGS_DIR / "data_audit.log"


def write_log(lines: list[str]) -> None:
    AUDIT_LOG.parent.mkdir(parents=True, exist_ok=True)
    AUDIT_LOG.write_text("\n".join(lines) + "\n", encoding="utf-8")


def get_excel_headers(path: Path, sheet_name: str) -> list[object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet_name]
        return [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    finally:
        workbook.close()


def duplicate_items(values: list[object]) -> list[object]:
    counts = Counter(values)
    return [value for value, count in counts.items() if count > 1]


def validate_source(log_lines: list[str]) -> pd.DataFrame:
    errors: list[str] = []

    if not INPUT_FILE.exists():
        errors.append(f"Input file not found: {relative_path(INPUT_FILE)}")
        write_log(log_lines + [f"ERROR: {errors[-1]}"])
        raise FileNotFoundError(errors[-1])

    excel_file = pd.ExcelFile(INPUT_FILE)
    log_lines.append(f"Workbook sheets: {excel_file.sheet_names}")
    if SHEET_NAME not in excel_file.sheet_names:
        errors.append(f"Sheet '{SHEET_NAME}' not found in {relative_path(INPUT_FILE)}")

    headers = get_excel_headers(INPUT_FILE, SHEET_NAME)
    duplicates = duplicate_items(headers)
    if duplicates:
        errors.append(f"Duplicate columns in Excel header: {duplicates}")

    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
    missing_columns = [column for column in EXPECTED_COLUMNS if column not in df.columns]
    extra_columns = [column for column in df.columns if column not in EXPECTED_COLUMNS]

    if missing_columns:
        errors.append(f"Missing expected columns: {missing_columns}")
    if extra_columns:
        log_lines.append(f"WARNING: Extra columns ignored: {extra_columns}")

    if errors:
        for error in errors:
            log_lines.append(f"ERROR: {error}")
        write_log(log_lines)
        raise ValueError("; ".join(errors))

    return df[EXPECTED_COLUMNS].copy()


def coerce_and_check_types(df: pd.DataFrame, log_lines: list[str]) -> pd.DataFrame:
    errors: list[str] = []
    working = df.copy()

    converted_dates = pd.to_datetime(working[DATE_COL], errors="coerce")
    invalid_dates = working.loc[converted_dates.isna(), DATE_COL]
    if not invalid_dates.empty:
        errors.append(
            f"Date conversion failed for {len(invalid_dates)} rows: "
            f"{invalid_dates.head(10).tolist()}"
        )
    working[DATE_COL] = converted_dates

    for column in NUMERIC_COLUMNS:
        original = working[column]
        converted = pd.to_numeric(original, errors="coerce")
        invalid_mask = original.notna() & converted.isna()
        if invalid_mask.any():
            examples = original.loc[invalid_mask].head(10).tolist()
            errors.append(
                f"Numeric conversion failed for '{column}' in "
                f"{int(invalid_mask.sum())} rows: {examples}"
            )
        working[column] = converted

    if errors:
        for error in errors:
            log_lines.append(f"ERROR: {error}")
        write_log(log_lines)
        raise ValueError("; ".join(errors))

    return working


def audit_series(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for column in NUMERIC_COLUMNS:
        series = df[column]
        finite_series = series.replace([np.inf, -np.inf], np.nan)
        valid = finite_series.dropna()

        if valid.empty:
            min_value = max_value = mean_value = median_value = std_value = np.nan
            min_date = max_date = ""
        else:
            min_idx = valid.idxmin()
            max_idx = valid.idxmax()
            min_value = valid.loc[min_idx]
            max_value = valid.loc[max_idx]
            min_date = format_date(df.loc[min_idx, DATE_COL])
            max_date = format_date(df.loc[max_idx, DATE_COL])
            mean_value = valid.mean()
            median_value = valid.median()
            std_value = valid.std(ddof=1)

        rows.append(
            {
                "variable": column,
                "observations": int(series.notna().sum()),
                "first_observation": format_date(df[DATE_COL].min()),
                "last_observation": format_date(df[DATE_COL].max()),
                "missing_values": int(series.isna().sum()),
                "zero_values": int((series == 0).sum()),
                "negative_values": int((series < 0).sum()),
                "minimum": min_value,
                "minimum_date": min_date,
                "maximum": max_value,
                "maximum_date": max_date,
                "mean": mean_value,
                "median": median_value,
                "standard_deviation": std_value,
            }
        )

    return pd.DataFrame(rows)


def missing_value_report(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    total_rows = len(df)
    for column in NUMERIC_COLUMNS:
        missing_mask = df[column].isna()
        missing_dates = df.loc[missing_mask, DATE_COL]
        rows.append(
            {
                "variable": column,
                "missing_values": int(missing_mask.sum()),
                "missing_share": float(missing_mask.mean()) if total_rows else np.nan,
                "first_missing_date": format_date(missing_dates.min()) if not missing_dates.empty else "",
                "last_missing_date": format_date(missing_dates.max()) if not missing_dates.empty else "",
            }
        )
    return pd.DataFrame(rows)


def check_calendar(df: pd.DataFrame, log_lines: list[str]) -> pd.DataFrame:
    was_ordered = df[DATE_COL].is_monotonic_increasing
    working = df.sort_values(DATE_COL).reset_index(drop=True)
    if not was_ordered:
        log_lines.append("WARNING: Dates were not ordered chronologically; data were sorted.")
    else:
        log_lines.append("Dates are ordered chronologically.")

    exact_duplicate_dates = working[DATE_COL].duplicated(keep=False)
    if exact_duplicate_dates.any():
        duplicate_dates = sorted(working.loc[exact_duplicate_dates, DATE_COL].dt.strftime("%Y-%m-%d").unique())
        raise ValueError(f"Duplicate dates found: {duplicate_dates}")

    working[DATE_COL] = to_month_end(working[DATE_COL])

    duplicate_months = working[DATE_COL].duplicated(keep=False)
    if duplicate_months.any():
        duplicate_values = sorted(working.loc[duplicate_months, DATE_COL].dt.strftime("%Y-%m-%d").unique())
        raise ValueError(f"Duplicate months found after month-end conversion: {duplicate_values}")

    missing = missing_months(working[DATE_COL])
    if missing:
        missing_text = [format_date(value) for value in missing]
        raise ValueError(f"Missing months inside sample: {missing_text}")

    log_lines.append("No duplicate dates, duplicate months, or internal missing months found.")
    return working


def build_common_sample(df: pd.DataFrame, log_lines: list[str]) -> pd.DataFrame:
    finite_values = finite_numeric_frame(df, NUMERIC_COLUMNS)
    complete_mask = finite_values.notna().all(axis=1)

    if complete_mask.all():
        clean = df.copy()
        log_lines.append("No rows removed for the common complete sample.")
    else:
        complete_positions = np.flatnonzero(complete_mask.to_numpy())
        if len(complete_positions) == 0:
            raise ValueError("No row has complete finite observations for all series.")

        first_position = int(complete_positions[0])
        last_position = int(complete_positions[-1])
        interior_mask = complete_mask.iloc[first_position : last_position + 1]
        if not interior_mask.all():
            bad_dates = df.loc[~complete_mask, DATE_COL].dt.strftime("%Y-%m-%d").tolist()
            raise ValueError(
                "Internal missing or infinite values prevent a consecutive common sample "
                f"without interpolation. Dates affected: {bad_dates}"
            )

        clean = df.iloc[first_position : last_position + 1].copy()
        log_lines.append(
            "Trimmed leading/trailing incomplete rows to keep the common sample: "
            f"{len(df) - len(clean)} rows removed."
        )

    missing_after_trim = missing_months(clean[DATE_COL])
    if missing_after_trim:
        missing_text = [format_date(value) for value in missing_after_trim]
        raise ValueError(f"Common sample is not consecutive monthly: {missing_text}")

    return clean[EXPECTED_COLUMNS].reset_index(drop=True)


def append_value_warnings(df: pd.DataFrame, log_lines: list[str]) -> None:
    missing_total = int(df[NUMERIC_COLUMNS].isna().sum().sum())
    infinite_total = int(np.isinf(df[NUMERIC_COLUMNS]).sum().sum())
    zero_total = int((df[NUMERIC_COLUMNS] == 0).sum().sum())
    negative_total = int((df[NUMERIC_COLUMNS] < 0).sum().sum())

    log_lines.append(f"Missing numeric values: {missing_total}")
    log_lines.append(f"Infinite numeric values: {infinite_total}")
    log_lines.append(f"Zero numeric values: {zero_total}")
    log_lines.append(f"Negative numeric values: {negative_total}")

    if missing_total:
        log_lines.append("WARNING: Missing values exist in the raw data.")
    if infinite_total:
        log_lines.append("WARNING: Infinite values exist in the raw data.")
    if zero_total:
        log_lines.append("WARNING: Zero values exist in the raw data.")
    if negative_total:
        log_lines.append("WARNING: Negative values exist in the raw data.")


def main() -> None:
    ensure_directories()
    log_lines = [
        "Data audit started.",
        f"Input workbook: {relative_path(INPUT_FILE)}",
        f"Input sheet: {SHEET_NAME}",
    ]

    try:
        raw_df = validate_source(log_lines)
        typed_df = coerce_and_check_types(raw_df, log_lines)
        monthly_df = check_calendar(typed_df, log_lines)

        audit_df = audit_series(monthly_df)
        missing_df = missing_value_report(monthly_df)
        save_csv(audit_df, AUDIT_TABLE)
        save_csv(missing_df, MISSING_TABLE)

        append_value_warnings(monthly_df, log_lines)
        clean_df = build_common_sample(monthly_df, log_lines)
        save_csv(clean_df, DATA_CLEAN_RAW)

        log_lines.append(f"Raw rows audited: {len(monthly_df)}")
        log_lines.append(f"Clean common-sample rows saved: {len(clean_df)}")
        log_lines.append(f"Sample period: {format_date(clean_df[DATE_COL].min())} to {format_date(clean_df[DATE_COL].max())}")
        log_lines.append("Data audit completed successfully.")
        write_log(log_lines)

        print("Data audit completed.")
        print(f"Saved {relative_path(AUDIT_TABLE)}")
        print(f"Saved {relative_path(MISSING_TABLE)}")
        print(f"Saved {relative_path(DATA_CLEAN_RAW)}")
    except Exception as exc:
        log_lines.append(f"ERROR: {exc}")
        write_log(log_lines)
        raise


if __name__ == "__main__":
    main()
